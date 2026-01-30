from __future__ import annotations

import os
import re
from datetime import datetime
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
import openpyxl

from excel_exporter import RmaEntry, export_to_excel, summarize_entries


MESES = [
    "JANEIRO",
    "FEVEREIRO",
    "MARÇO",
    "ABRIL",
    "MAIO",
    "JUNHO",
    "JULHO",
    "AGOSTO",
    "SETEMBRO",
    "OUTUBRO",
    "NOVEMBRO",
    "DEZEMBRO",
]

PALETTE = [
    "#FFD966",
    "#F4B183",
    "#C6E0B4",
    "#9DC3E6",
    "#D9D2E9",
    "#F8CBAD",
    "#A9D18E",
    "#8FAADC",
    "#E2EFDA",
    "#C9C9C9",
]


class RmaApp(tk.Tk):
    def __init__(self) -> None:
        super().__init__()

        now = datetime.now()

        self.title("Gerador de Planilha RMA")
        self.geometry("1300x760")
        self.minsize(1100, 650)

        self.planilha_titulo_var = tk.StringVar(value=f"{now:%d/%m/%Y}(Atualizada) Planilha RMA")
        self.periodo_mes_var = tk.StringVar(value=MESES[now.month - 1])
        self.periodo_ano_var = tk.StringVar(value=str(now.year))
        self.abrir_ao_exportar_var = tk.BooleanVar(value=True)

        self.vars: dict[str, tk.StringVar] = {
            "recebimento": tk.StringVar(value=now.strftime("%d/%m/%Y")),
            "cliente": tk.StringVar(),
            "nf": tk.StringVar(),
            "os": tk.StringVar(),
            "triagem": tk.StringVar(),
            "produto_enviado": tk.StringVar(),
            "und": tk.StringVar(),
            "plataforma": tk.StringVar(),
            "codigo": tk.StringVar(),
            "numero_serie": tk.StringVar(),
            "status": tk.StringVar(),
            "configuracao_avaria": tk.StringVar(),
            "pedido_marketplace": tk.StringVar(),
        }

        self.entry_counter = 0
        self.entry_by_id: dict[str, RmaEntry] = {}
        self.editing_id: str | None = None

        self.laudo_text: tk.Text | None = None
        self.add_update_button: ttk.Button | None = None

        self.chart_fig: Figure | None = None
        self.chart_ax = None
        self.chart_canvas: FigureCanvasTkAgg | None = None

        self.tree: ttk.Treeview | None = None
        self.pieces_tree: ttk.Treeview | None = None
        self.reasons_tree: ttk.Treeview | None = None

        self._build_ui()
        self._refresh_summaries()

    def _build_ui(self) -> None:
        main = ttk.Frame(self)
        main.pack(fill="both", expand=True)

        main.columnconfigure(0, weight=3)
        main.columnconfigure(1, weight=2)
        main.rowconfigure(0, weight=1)

        left = ttk.Frame(main)
        left.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
        left.columnconfigure(0, weight=1)
        left.rowconfigure(3, weight=1)

        right = ttk.Frame(main)
        right.grid(row=0, column=1, sticky="nsew", padx=(0, 10), pady=10)
        right.columnconfigure(0, weight=1)
        right.rowconfigure(0, weight=1)

        meta = ttk.LabelFrame(left, text="Configuração")
        meta.grid(row=0, column=0, sticky="ew", padx=0, pady=(0, 10))
        meta.columnconfigure(1, weight=1)

        ttk.Label(meta, text="Título da planilha").grid(row=0, column=0, sticky="w", padx=6, pady=4)
        ttk.Entry(meta, textvariable=self.planilha_titulo_var).grid(
            row=0, column=1, columnspan=5, sticky="ew", padx=6, pady=4
        )

        ttk.Label(meta, text="Mês").grid(row=1, column=0, sticky="w", padx=6, pady=4)
        ttk.Combobox(meta, values=MESES, textvariable=self.periodo_mes_var, state="readonly", width=16).grid(
            row=1, column=1, sticky="w", padx=6, pady=4
        )

        ttk.Label(meta, text="Ano").grid(row=1, column=2, sticky="w", padx=6, pady=4)
        ttk.Entry(meta, textvariable=self.periodo_ano_var, width=8).grid(row=1, column=3, sticky="w", padx=6, pady=4)

        ttk.Checkbutton(meta, text="Abrir Excel após exportar", variable=self.abrir_ao_exportar_var).grid(
            row=1, column=4, sticky="w", padx=6, pady=4
        )

        ttk.Button(meta, text="Selecionar Planilha", command=self._import_excel).grid(
            row=1, column=5, sticky="e", padx=6, pady=4
        )

        ttk.Button(meta, text="Exportar Excel", command=self._export_excel).grid(
            row=1, column=6, sticky="e", padx=6, pady=4
        )

        form = ttk.LabelFrame(left, text="Cadastro")
        form.grid(row=1, column=0, sticky="ew", padx=0, pady=(0, 10))

        form.columnconfigure(1, weight=1)
        form.columnconfigure(3, weight=1)

        left_fields = [
            ("RECEBIMENTO", "recebimento"),
            ("Cliente", "cliente"),
            ("NF", "nf"),
            ("OS", "os"),
            ("Triagem", "triagem"),
            ("Produto enviado", "produto_enviado"),
            ("UND", "und"),
        ]
        right_fields = [
            ("Plataforma", "plataforma"),
            ("Código", "codigo"),
            ("Numero de serie", "numero_serie"),
            ("Status", "status"),
            ("Configuração/Avaria", "configuracao_avaria"),
            ("Pedido Marketplace", "pedido_marketplace"),
        ]

        for row, (label, key) in enumerate(left_fields):
            self._add_labeled_entry(form, label, self.vars[key], row, 0)
        for row, (label, key) in enumerate(right_fields):
            self._add_labeled_entry(form, label, self.vars[key], row, 2)

        ttk.Label(form, text="LAUDO TÉCNICO").grid(row=7, column=0, sticky="w", padx=6, pady=(8, 4))

        laudo_frame = ttk.Frame(form)
        laudo_frame.grid(row=8, column=0, columnspan=4, sticky="ew", padx=6, pady=(0, 8))
        laudo_frame.columnconfigure(0, weight=1)

        self.laudo_text = tk.Text(laudo_frame, height=4, wrap="word")
        laudo_scroll = ttk.Scrollbar(laudo_frame, orient="vertical", command=self.laudo_text.yview)
        self.laudo_text.configure(yscrollcommand=laudo_scroll.set)

        self.laudo_text.grid(row=0, column=0, sticky="ew")
        laudo_scroll.grid(row=0, column=1, sticky="ns")

        actions = ttk.Frame(left)
        actions.grid(row=2, column=0, sticky="ew", padx=0, pady=(0, 10))
        actions.columnconfigure(5, weight=1)

        self.add_update_button = ttk.Button(actions, text="Adicionar", command=self._add_or_update_entry)
        self.add_update_button.grid(row=0, column=0, padx=4, pady=4, sticky="w")
        ttk.Button(actions, text="Limpar", command=self._clear_form).grid(row=0, column=1, padx=4, pady=4, sticky="w")
        ttk.Button(actions, text="Editar selecionado", command=self._edit_selected).grid(
            row=0, column=2, padx=4, pady=4, sticky="w"
        )
        ttk.Button(actions, text="Excluir selecionado", command=self._delete_selected).grid(
            row=0, column=3, padx=4, pady=4, sticky="w"
        )
        ttk.Button(actions, text="Colar Dados", command=self._paste_data).grid(
            row=0, column=4, padx=4, pady=4, sticky="w"
        )

        table_frame = ttk.LabelFrame(left, text="Registros")
        table_frame.grid(row=3, column=0, sticky="nsew")
        table_frame.columnconfigure(0, weight=1)
        table_frame.rowconfigure(0, weight=1)

        cols = [
            "RECEBIMENTO",
            "Cliente",
            "NF",
            "OS",
            "Triagem",
            "Produto enviado",
            "UND",
            "Plataforma",
            "Código",
            "Numero de serie",
            "Status",
            "Configuração/Avaria",
            "Pedido Marketplace",
            "LAUDO TÉCNICO",
        ]

        self.tree = ttk.Treeview(table_frame, columns=cols, show="headings", selectmode="extended")
        for c in cols:
            self.tree.heading(c, text=c)
            self.tree.column(c, width=110, stretch=True)

        self.tree.column("RECEBIMENTO", width=110, stretch=False)
        self.tree.column("Cliente", width=160, stretch=True)
        self.tree.column("Produto enviado", width=180, stretch=True)
        self.tree.column("Configuração/Avaria", width=220, stretch=True)
        self.tree.column("LAUDO TÉCNICO", width=260, stretch=True)

        xscroll = ttk.Scrollbar(table_frame, orient="horizontal", command=self.tree.xview)
        yscroll = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(xscrollcommand=xscroll.set, yscrollcommand=yscroll.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        yscroll.grid(row=0, column=1, sticky="ns")
        xscroll.grid(row=1, column=0, sticky="ew")

        self.tree.bind("<Double-1>", lambda _evt: self._edit_selected())

        notebook = ttk.Notebook(right)
        notebook.grid(row=0, column=0, sticky="nsew")

        chart_tab = ttk.Frame(notebook)
        summary_tab = ttk.Frame(notebook)
        notebook.add(chart_tab, text="Gráfico")
        notebook.add(summary_tab, text="Resumo")

        chart_tab.columnconfigure(0, weight=1)
        chart_tab.rowconfigure(0, weight=1)

        self.chart_fig = Figure(figsize=(6, 4), dpi=100)
        self.chart_ax = self.chart_fig.add_subplot(111)
        self.chart_canvas = FigureCanvasTkAgg(self.chart_fig, master=chart_tab)
        self.chart_canvas.get_tk_widget().grid(row=0, column=0, sticky="nsew", padx=8, pady=8)

        summary_tab.columnconfigure(0, weight=1)
        summary_tab.rowconfigure(0, weight=1)
        summary_tab.rowconfigure(1, weight=1)

        pieces_frame = ttk.LabelFrame(summary_tab, text="Peças defeituosas")
        pieces_frame.grid(row=0, column=0, sticky="nsew", padx=8, pady=(8, 6))
        pieces_frame.columnconfigure(0, weight=1)
        pieces_frame.rowconfigure(0, weight=1)

        self.pieces_tree = ttk.Treeview(pieces_frame, columns=["Peça", "Qtd"], show="headings")
        self.pieces_tree.heading("Peça", text="Peça")
        self.pieces_tree.heading("Qtd", text="Qtd")
        self.pieces_tree.column("Peça", width=260, stretch=True)
        self.pieces_tree.column("Qtd", width=60, stretch=False, anchor="center")

        pieces_scroll = ttk.Scrollbar(pieces_frame, orient="vertical", command=self.pieces_tree.yview)
        self.pieces_tree.configure(yscrollcommand=pieces_scroll.set)
        self.pieces_tree.grid(row=0, column=0, sticky="nsew")
        pieces_scroll.grid(row=0, column=1, sticky="ns")

        reasons_frame = ttk.LabelFrame(summary_tab, text="Motivos defeituosos")
        reasons_frame.grid(row=1, column=0, sticky="nsew", padx=8, pady=(6, 8))
        reasons_frame.columnconfigure(0, weight=1)
        reasons_frame.rowconfigure(0, weight=1)

        self.reasons_tree = ttk.Treeview(reasons_frame, columns=["Motivo", "Qtd"], show="headings")
        self.reasons_tree.heading("Motivo", text="Motivo")
        self.reasons_tree.heading("Qtd", text="Qtd")
        self.reasons_tree.column("Motivo", width=260, stretch=True)
        self.reasons_tree.column("Qtd", width=60, stretch=False, anchor="center")

        reasons_scroll = ttk.Scrollbar(reasons_frame, orient="vertical", command=self.reasons_tree.yview)
        self.reasons_tree.configure(yscrollcommand=reasons_scroll.set)
        self.reasons_tree.grid(row=0, column=0, sticky="nsew")
        reasons_scroll.grid(row=0, column=1, sticky="ns")

    def _add_labeled_entry(self, parent: ttk.Widget, label: str, var: tk.StringVar, row: int, col: int) -> None:
        ttk.Label(parent, text=label).grid(row=row, column=col, sticky="w", padx=6, pady=2)
        ttk.Entry(parent, textvariable=var).grid(row=row, column=col + 1, sticky="ew", padx=6, pady=2)

    def _entry_to_values(self, e: RmaEntry) -> list[str]:
        return [
            e.recebimento,
            e.cliente,
            e.nf,
            e.os,
            e.triagem,
            e.produto_enviado,
            e.und,
            e.plataforma,
            e.codigo,
            e.numero_serie,
            e.status,
            e.configuracao_avaria,
            e.pedido_marketplace,
            e.laudo_tecnico,
        ]

    def _get_entries_in_display_order(self) -> list[RmaEntry]:
        if self.tree is None:
            return []
        out: list[RmaEntry] = []
        for iid in self.tree.get_children(""):
            e = self.entry_by_id.get(iid)
            if e is not None:
                out.append(e)
        return out

    def _collect_form_entry(self) -> RmaEntry:
        laudo = ""
        if self.laudo_text is not None:
            laudo = self.laudo_text.get("1.0", "end").rstrip("\n")

        return RmaEntry(
            recebimento=self.vars["recebimento"].get().strip(),
            cliente=self.vars["cliente"].get().strip(),
            nf=self.vars["nf"].get().strip(),
            os=self.vars["os"].get().strip(),
            triagem=self.vars["triagem"].get().strip(),
            produto_enviado=self.vars["produto_enviado"].get().strip(),
            und=self.vars["und"].get().strip(),
            plataforma=self.vars["plataforma"].get().strip(),
            codigo=self.vars["codigo"].get().strip(),
            numero_serie=self.vars["numero_serie"].get().strip(),
            status=self.vars["status"].get().strip(),
            configuracao_avaria=self.vars["configuracao_avaria"].get().strip(),
            pedido_marketplace=self.vars["pedido_marketplace"].get().strip(),
            laudo_tecnico=laudo.strip(),
        )

    def _add_or_update_entry(self) -> None:
        if self.tree is None:
            return

        entry = self._collect_form_entry()

        if self.editing_id is not None:
            iid = self.editing_id
            self.entry_by_id[iid] = entry
            self.tree.item(iid, values=self._entry_to_values(entry))
            self.editing_id = None
            if self.add_update_button is not None:
                self.add_update_button.configure(text="Adicionar")
        else:
            self.entry_counter += 1
            iid = str(self.entry_counter)
            self.entry_by_id[iid] = entry
            self.tree.insert("", "end", iid=iid, values=self._entry_to_values(entry))

        self._clear_form(keep_recebimento=True)
        self._refresh_summaries()

    def _clear_form(self, *, keep_recebimento: bool = False) -> None:
        for k, var in self.vars.items():
            if keep_recebimento and k == "recebimento":
                continue
            var.set("")

        if self.laudo_text is not None:
            self.laudo_text.delete("1.0", "end")

        self.editing_id = None
        if self.add_update_button is not None:
            self.add_update_button.configure(text="Adicionar")

    def _edit_selected(self) -> None:
        if self.tree is None:
            return

        sel = self.tree.selection()
        if not sel:
            messagebox.showwarning("Editar", "Selecione um registro para editar.")
            return

        iid = sel[0]
        entry = self.entry_by_id.get(iid)
        if entry is None:
            return

        self.vars["recebimento"].set(entry.recebimento)
        self.vars["cliente"].set(entry.cliente)
        self.vars["nf"].set(entry.nf)
        self.vars["os"].set(entry.os)
        self.vars["triagem"].set(entry.triagem)
        self.vars["produto_enviado"].set(entry.produto_enviado)
        self.vars["und"].set(entry.und)
        self.vars["plataforma"].set(entry.plataforma)
        self.vars["codigo"].set(entry.codigo)
        self.vars["numero_serie"].set(entry.numero_serie)
        self.vars["status"].set(entry.status)
        self.vars["configuracao_avaria"].set(entry.configuracao_avaria)
        self.vars["pedido_marketplace"].set(entry.pedido_marketplace)

        if self.laudo_text is not None:
            self.laudo_text.delete("1.0", "end")
            self.laudo_text.insert("1.0", entry.laudo_tecnico)

        self.editing_id = iid
        if self.add_update_button is not None:
            self.add_update_button.configure(text="Atualizar")

    def _delete_selected(self) -> None:
        if self.tree is None:
            return

        sel = list(self.tree.selection())
        if not sel:
            messagebox.showwarning("Excluir", "Selecione um ou mais registros para excluir.")
            return

        if not messagebox.askyesno("Excluir", f"Excluir {len(sel)} registro(s)?"):
            return

        for iid in sel:
            self.tree.delete(iid)
            self.entry_by_id.pop(iid, None)
            if self.editing_id == iid:
                self.editing_id = None

        if self.add_update_button is not None:
            self.add_update_button.configure(text="Adicionar")

        self._refresh_summaries()

    def _refresh_summaries(self) -> None:
        entries = self._get_entries_in_display_order()
        pieces_sorted, reasons_sorted = summarize_entries(entries)

        if self.pieces_tree is not None:
            for iid in self.pieces_tree.get_children(""):
                self.pieces_tree.delete(iid)
            for name, qty in pieces_sorted:
                self.pieces_tree.insert("", "end", values=[name, qty])
            if pieces_sorted:
                self.pieces_tree.insert("", "end", values=["TOTAL", sum(q for _, q in pieces_sorted)])

        if self.reasons_tree is not None:
            for iid in self.reasons_tree.get_children(""):
                self.reasons_tree.delete(iid)
            for name, qty in reasons_sorted:
                self.reasons_tree.insert("", "end", values=[name, qty])
            if reasons_sorted:
                self.reasons_tree.insert("", "end", values=["TOTAL", sum(q for _, q in reasons_sorted)])

        if self.chart_ax is None or self.chart_canvas is None:
            return

        self.chart_ax.clear()

        if not pieces_sorted:
            self.chart_ax.text(0.5, 0.5, "Sem dados", ha="center", va="center")
            self.chart_ax.set_axis_off()
            self.chart_canvas.draw_idle()
            return

        labels = [name for name, _qty in pieces_sorted]
        values = [qty for _name, qty in pieces_sorted]
        colors = [PALETTE[i % len(PALETTE)] for i in range(len(values))]

        wedges, _texts, autotexts = self.chart_ax.pie(
            values,
            labels=None,
            autopct=lambda pct: f"{pct:.0f}%",
            startangle=90,
            colors=colors,
            wedgeprops={"width": 0.45, "edgecolor": "white"},
        )
        for t in autotexts:
            t.set_color("white")
            t.set_fontsize(9)

        self.chart_ax.set_aspect("equal")
        self.chart_ax.set_title(
            f"PEÇAS DEFEITUOSAS\n{self.periodo_mes_var.get()} - {self.periodo_ano_var.get()}"
        )
        self.chart_ax.legend(
            wedges,
            labels,
            loc="upper center",
            bbox_to_anchor=(0.5, 1.02),
            ncol=2,
            frameon=False,
            fontsize=8,
        )
        self.chart_canvas.draw_idle()

    def _export_excel(self) -> None:
        entries = self._get_entries_in_display_order()
        if not entries:
            messagebox.showwarning("Exportar", "Adicione pelo menos um registro antes de exportar.")
            return

        now = datetime.now()
        default_name = f"Planilha_RMA_{now:%Y-%m-%d}.xlsx"

        file_name = filedialog.asksaveasfilename(
            title="Salvar planilha",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=default_name,
        )
        if not file_name:
            return

        try:
            path = export_to_excel(
                entries,
                file_name,
                title=self.planilha_titulo_var.get().strip() or "Planilha RMA",
                periodo_mes=self.periodo_mes_var.get().strip() or "",
                periodo_ano=self.periodo_ano_var.get().strip() or "",
            )
        except Exception as e:
            messagebox.showerror("Exportar", f"Falha ao gerar o Excel:\n{e}")
            return

        messagebox.showinfo("Exportar", f"Planilha gerada com sucesso:\n{path}")

        if self.abrir_ao_exportar_var.get():
            try:
                os.startfile(str(path))  # type: ignore[attr-defined]
            except Exception:
                pass


    def _import_excel(self) -> None:
        file_path = filedialog.askopenfilename(
            title="Selecionar planilha existente",
            filetypes=[("Excel", "*.xlsx")],
        )
        if not file_path:
            return

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
        except Exception as e:
            messagebox.showerror("Importar", f"Erro ao abrir o arquivo:\n{e}")
            return

        if "RMA" not in wb.sheetnames:
            messagebox.showerror("Importar", "A planilha não contém a aba 'RMA'.")
            wb.close()
            return

        ws = wb["RMA"]
        imported_count = 0

        for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
            if not row or all(cell is None or str(cell).strip() == "" for cell in row):
                continue

            def safe(val: object) -> str:
                return str(val).strip() if val is not None else ""

            entry = RmaEntry(
                recebimento=safe(row[0]) if len(row) > 0 else "",
                cliente=safe(row[1]) if len(row) > 1 else "",
                nf=safe(row[2]) if len(row) > 2 else "",
                os=safe(row[3]) if len(row) > 3 else "",
                triagem=safe(row[4]) if len(row) > 4 else "",
                produto_enviado=safe(row[5]) if len(row) > 5 else "",
                und=safe(row[6]) if len(row) > 6 else "",
                plataforma=safe(row[7]) if len(row) > 7 else "",
                codigo=safe(row[8]) if len(row) > 8 else "",
                numero_serie=safe(row[9]) if len(row) > 9 else "",
                status=safe(row[10]) if len(row) > 10 else "",
                configuracao_avaria=safe(row[11]) if len(row) > 11 else "",
                pedido_marketplace=safe(row[12]) if len(row) > 12 else "",
                laudo_tecnico=safe(row[13]) if len(row) > 13 else "",
            )

            self.entry_counter += 1
            iid = str(self.entry_counter)
            self.entry_by_id[iid] = entry
            if self.tree is not None:
                self.tree.insert("", "end", iid=iid, values=self._entry_to_values(entry))
            imported_count += 1

        wb.close()
        self._refresh_summaries()
        messagebox.showinfo("Importar", f"{imported_count} registro(s) importado(s) com sucesso!")

    def _paste_data(self) -> None:
        try:
            clipboard = self.clipboard_get()
        except tk.TclError:
            messagebox.showwarning("Colar Dados", "Área de transferência vazia ou sem texto.")
            return

        if not clipboard.strip():
            messagebox.showwarning("Colar Dados", "Nenhum dado para colar.")
            return

        lines = clipboard.strip().split("\n")
        imported_count = 0

        for line in lines:
            line = line.strip()
            if not line:
                continue

            parts = re.split(r"\t", line)
            if len(parts) < 2:
                parts = re.split(r"\s{2,}", line)

            def safe(idx: int) -> str:
                return parts[idx].strip() if idx < len(parts) else ""

            entry = RmaEntry(
                recebimento=safe(0),
                cliente=safe(1),
                nf=safe(2),
                os=safe(3),
                triagem=safe(4),
                produto_enviado=safe(5),
                und=safe(6),
                plataforma=safe(7),
                codigo=safe(8),
                numero_serie=safe(9),
                status=safe(10),
                configuracao_avaria=safe(11),
                pedido_marketplace=safe(12),
                laudo_tecnico=safe(13),
            )

            self.entry_counter += 1
            iid = str(self.entry_counter)
            self.entry_by_id[iid] = entry
            if self.tree is not None:
                self.tree.insert("", "end", iid=iid, values=self._entry_to_values(entry))
            imported_count += 1

        self._refresh_summaries()
        if imported_count > 0:
            messagebox.showinfo("Colar Dados", f"{imported_count} registro(s) adicionado(s)!")
        else:
            messagebox.showwarning("Colar Dados", "Nenhum dado válido encontrado.")


def main() -> None:
    app = RmaApp()
    app.mainloop()


if __name__ == "__main__":
    main()
